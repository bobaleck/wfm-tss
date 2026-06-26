from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date, timedelta, time as _time, datetime as _datetime

from app.api.deps import get_current_user, check_project_access, require_manager
from app.core.database import get_db
from app.models.audit import IntegrationSettings, QueueSetting, StatusConfig, CustomerDemand
from app.models.employee import Employee
from app.services.status_classification import build_status_sets
import app.services.naumen_db as naumen

router = APIRouter()

# Кэш карты «оператор → очереди» (по обработанным звонкам за неделю). Мониторинг
# опрашивает current-operators каждые 5 секунд, а очереди оператора меняются
# медленно — поэтому держим результат до 5 минут, чтобы не сканировать Naumen
# на каждый опрос. Ключ — partner_uuid; значение — (timestamp, map).
import time as _time_mod
_QUEUES_MAP_TTL = 300
_queues_map_cache: dict[str, tuple[float, dict]] = {}


def _operator_queues_cached(partner_uuid: str, overrides: Optional[dict]) -> dict:
    now = _time_mod.time()
    hit = _queues_map_cache.get(partner_uuid)
    if hit and now - hit[0] < _QUEUES_MAP_TTL:
        return hit[1]
    data = naumen.get_operator_queues_map(partner_uuid, 7, overrides)
    _queues_map_cache[partner_uuid] = (now, data)
    return data


# Аналогичный кэш для карты «оператор → исходящие проекты (линии)».
_outbound_projects_cache: dict[str, tuple[float, dict]] = {}


def _operator_outbound_projects_cached(partner_uuid: str, overrides: Optional[dict]) -> dict:
    now = _time_mod.time()
    hit = _outbound_projects_cache.get(partner_uuid)
    if hit and now - hit[0] < _QUEUES_MAP_TTL:
        return hit[1]
    data = naumen.get_operator_outbound_projects_map(partner_uuid, 7, overrides)
    _outbound_projects_cache[partner_uuid] = (now, data)
    return data


# Кэш СПИСКА исходящих подпроектов (мало меняется, а запрашивается со многих
# страниц — дашборд/мониторинг/смены/настройки). Кэшируем naumen-часть, чтобы не
# открывать соединение к Naumen на каждый запрос; флаги Вход/Исход/Скрыть из
# нашей БД применяем поверх кэша на каждый вызов.
_outbound_list_cache: dict[str, tuple[float, list]] = {}


def _outbound_projects_list_cached(partner_uuid: str, overrides: Optional[dict]) -> list:
    now = _time_mod.time()
    hit = _outbound_list_cache.get(partner_uuid)
    if hit and now - hit[0] < _QUEUES_MAP_TTL:
        return [dict(p) for p in hit[1]]   # копия, чтобы не мутировать кэш флагами
    data = naumen.get_outbound_projects(partner_uuid, overrides)
    _outbound_list_cache[partner_uuid] = (now, data)
    return [dict(p) for p in data]


def _exclusive_end(end: date) -> str:
    """Фронтенд присылает диапазон ВКЛЮЧИТЕЛЬНО [begin, end] (для одного дня
    begin == end). В SQL верхняя граница строгая (enqueued_time < end), поэтому
    добавляем один день — тогда и одиночный день («Вчера»), и последний день
    месяца/квартала корректно попадают в выборку."""
    return str(end + timedelta(days=1))


def _build_overrides(db: Session) -> Optional[dict]:
    s = db.query(IntegrationSettings).first()
    if s and s.db_host:
        return {
            "host": s.db_host,
            "database": s.db_name,
            "user": s.db_user,
            "password": s.db_password,
            "port": s.db_port,
        }
    return None


def _status_sets(db: Session, partner_uuid: str) -> tuple[list[str], list[str]]:
    """Рабочие/офлайн статусы для проекта — стандартные + индивидуальные настройки.
    Та же классификация используется и в Онлайн-мониторинге, и в истории смен."""
    configs = db.query(StatusConfig).filter(StatusConfig.project_uuid == partner_uuid).all()
    return build_status_sets(configs)


@router.get("/projects")
def get_projects(db: Session = Depends(get_db), _=Depends(get_current_user)):
    try:
        return {"data": naumen.get_projects(_build_overrides(db))}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/queues")
def get_queues(partner_uuid: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    check_project_access(partner_uuid, current_user, db)
    try:
        queues = naumen.get_queues(partner_uuid, _build_overrides(db))
        # Apply WFM overrides (target_sl, answer_sec) + направление/видимость очереди.
        overrides = {
            r.queue_name: r
            for r in db.query(QueueSetting).filter(QueueSetting.partner_uuid == partner_uuid).all()
        }
        for q in queues:
            ov = overrides.get(q.get("name"))
            # По умолчанию входящая очередь = «Вход», видима.
            q["show_in"] = bool(ov.show_in) if ov and ov.show_in is not None else True
            q["show_out"] = bool(ov.show_out) if ov and ov.show_out is not None else False
            q["hidden"] = bool(ov.hidden) if ov and ov.hidden is not None else False
            if ov:
                if ov.target_sl is not None:
                    q["target_sl"] = ov.target_sl
                if ov.answer_sec is not None:
                    q["answer_sec"] = ov.answer_sec
        return {"data": queues}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/workload")
def get_workload(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    interval: str = "hour",
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_project_access(partner_uuid, current_user, db)
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        data = naumen.get_workload(partner_uuid, str(begin), _exclusive_end(end), interval, _build_overrides(db))
        return {"data": data, "meta": {"begin": str(begin), "end": str(end), "interval": interval}}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/operator-load")
def get_operator_load(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_project_access(partner_uuid, current_user, db)
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        work_statuses, offline_statuses = _status_sets(db, partner_uuid)
        data = naumen.get_operator_load(
            partner_uuid, str(begin), _exclusive_end(end), work_statuses, offline_statuses, _build_overrides(db),
        )
        return {"data": data, "meta": {"begin": str(begin), "end": str(end)}}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/operator-load-by-queue")
def get_operator_load_by_queue_ep(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Нагрузка операторов в разрезе очередей (для раскрытия по очередям и по оператору)."""
    check_project_access(partner_uuid, current_user, db)
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        data = naumen.get_operator_load_by_queue(
            partner_uuid, str(begin), _exclusive_end(end), _build_overrides(db),
        )
        return {"data": data}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/status-summary")
def get_status_summary(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    try:
        data = naumen.get_status_summary(partner_uuid, str(begin), _exclusive_end(end), _build_overrides(db))
        return {"data": data}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/operator-sessions")
def get_operator_sessions(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    История сессий операторов из Naumen status_changes.
    Берёт логины сотрудников проекта из локальной БД, обогащает именами.
    """
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")

    # Логины сотрудников этого проекта из нашей БД
    employees = db.query(Employee).filter(
        Employee.project_uuid == partner_uuid,
        Employee.naumen_login.isnot(None),
    ).all()
    login_map = {e.naumen_login: e.full_name for e in employees}
    logins = list(login_map.keys())

    if not logins:
        return {"data": [], "meta": {"logins_found": 0}}

    try:
        work_statuses, offline_statuses = _status_sets(db, partner_uuid)
        rows = naumen.get_operator_sessions(
            logins, str(begin), _exclusive_end(end), work_statuses, offline_statuses, _build_overrides(db),
        )
        for r in rows:
            r["employee_name"] = login_map.get(r.get("login"), r.get("login"))
        return {"data": rows, "meta": {"logins_found": len(logins), "begin": str(begin), "end": str(end)}}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/operator-timeline")
def get_operator_timeline(
    login: str,
    work_date: Optional[str] = None,
    hours: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Временная линия статусов оператора — либо за календарный день (work_date,
    используется в Сменах), либо за скользящее окно в hours часов до текущего
    момента (используется в Онлайн-мониторинге)."""
    try:
        if hours is not None:
            data = naumen.get_operator_timeline_window(login, hours, _build_overrides(db))
        elif work_date is not None:
            data = naumen.get_operator_timeline(login, work_date, _build_overrides(db))
        else:
            raise HTTPException(400, detail="Укажите work_date или hours")
        return {"data": data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/actual-operators")
def get_actual_operators(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Среднее фактическое число операторов по часам суток за период."""
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    employees = db.query(Employee).filter(
        Employee.project_uuid == partner_uuid,
        Employee.naumen_login.isnot(None),
    ).all()
    logins = [e.naumen_login for e in employees]
    try:
        data = naumen.get_actual_operators_by_hour(logins, str(begin), _exclusive_end(end), _build_overrides(db))
        return {"data": data}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


def _parse_hour(v):
    """Извлекает час (0..23) из значения ячейки-заголовка (time/datetime/строка/int)."""
    if isinstance(v, _time):
        return v.hour
    if isinstance(v, _datetime):
        return v.hour
    if isinstance(v, int) and 0 <= v <= 23:
        return v
    if isinstance(v, str):
        s = v.strip()
        if ":" in s:
            try:
                return int(s.split(":")[0]) % 24
            except Exception:
                return None
    return None


def _parse_demand_date(v):
    if isinstance(v, _datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


@router.post("/customer-demand/upload")
async def upload_customer_demand(
    partner_uuid: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_manager),
):
    """Загрузка потребности «от заказчика» из Excel и сохранение на проект.
    Лист со строками-датами и колонками-часами (00:00..23:00); число дней любое."""
    import io
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, detail=f"Не удалось прочитать Excel: {e}")

    parsed = []  # (date, hour, required)
    for ws in wb.worksheets:
        hour_cols, header_row = {}, None
        for r in range(1, min(ws.max_row, 25) + 1):
            cols = {}
            for c in range(1, ws.max_column + 1):
                h = _parse_hour(ws.cell(r, c).value)
                if h is not None:
                    cols[c] = h
            if len(cols) >= 12:  # строка с ~24 часами = шапка
                header_row, hour_cols = r, cols
                break
        if not header_row:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            dt = None
            for c in range(1, 7):
                dt = _parse_demand_date(ws.cell(r, c).value)
                if dt:
                    break
            if not dt:
                continue
            for c, h in hour_cols.items():
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    parsed.append((dt, h, int(round(v))))
        if parsed:
            break  # берём первый лист с данными

    if not parsed:
        raise HTTPException(400, detail="В файле не найдены строки с датами и часовой потребностью")

    # Полная замена потребности проекта на загруженную
    db.query(CustomerDemand).filter(CustomerDemand.project_uuid == partner_uuid).delete()
    seen = set()
    for dt, h, req in parsed:
        key = (dt, h)
        if key in seen:
            continue
        seen.add(key)
        db.add(CustomerDemand(project_uuid=partner_uuid, demand_date=dt, hour=h, required=req))
    db.commit()
    dates = sorted({dt for dt, _, _ in parsed})
    return {"ok": True, "rows": len(seen), "days": len(dates),
            "date_from": str(dates[0]), "date_to": str(dates[-1])}


@router.get("/customer-demand/template.xlsx")
def customer_demand_template(_=Depends(get_current_user)):
    """Пустой Excel-шаблон потребности (текущий месяц, даты × 24 часа) для заполнения
    и последующей загрузки."""
    import io
    from datetime import date as _date, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = Workbook(); ws = wb.active; ws.title = "Потребность"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="434343")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    WEEKDAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    ws.cell(row=1, column=2, value="Дата")
    ws.cell(row=1, column=3, value="День")
    for h in range(24):
        ws.cell(row=1, column=5 + h, value=f"{h:02d}:00")
    for col in range(2, 5 + 24):
        cc = ws.cell(row=1, column=col)
        cc.fill = hdr_fill; cc.font = hdr_font; cc.alignment = center; cc.border = border

    today = _date.today()
    first = today.replace(day=1)
    d, r = first, 2
    while d.month == first.month:
        dc = ws.cell(row=r, column=2, value=d); dc.number_format = "DD.MM.YYYY"; dc.border = border
        wc = ws.cell(row=r, column=3, value=WEEKDAYS[d.weekday()]); wc.border = border
        for h in range(24):
            cc = ws.cell(row=r, column=5 + h); cc.border = border; cc.alignment = center
        d += timedelta(days=1); r += 1

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 6
    for h in range(24):
        ws.column_dimensions[get_column_letter(5 + h)].width = 6

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customer_demand_template.xlsx"'},
    )


@router.get("/customer-demand")
def get_customer_demand(
    partner_uuid: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    rows = db.query(CustomerDemand).filter(
        CustomerDemand.project_uuid == partner_uuid,
    ).order_by(CustomerDemand.demand_date, CustomerDemand.hour).all()
    return {"data": [
        {"demand_date": str(r.demand_date), "hour": r.hour, "required": r.required}
        for r in rows
    ]}


@router.get("/recent-stats")
def get_recent_stats_ep(
    partner_uuid: str,
    window_min: int = Query(1440, ge=1, le=1440),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Статистика за последние window_min минут (макс. 24ч=1440): по очередям и по
    (оператор, очередь). Для раздела «Мониторинг → Статистика» с ползунком окна."""
    check_project_access(partner_uuid, current_user, db)
    try:
        return naumen.get_recent_stats(partner_uuid, int(window_min), _build_overrides(db))
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/actual-operators-by-queue")
def get_actual_operators_by_queue_ep(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    queues: Optional[list[str]] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Факт операторов по часам с учётом очередей. Если queues заданы — union
    по выбранным очередям (мультиочередной оператор не задваивается). Доп. ключ
    by_queue — разрез по каждой очереди."""
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        union = naumen.get_actual_operators_union(
            partner_uuid, str(begin), _exclusive_end(end), queues or None, _build_overrides(db),
        )
        by_queue = naumen.get_actual_operators_by_queue(
            partner_uuid, str(begin), _exclusive_end(end), _build_overrides(db),
        )
        return {"data": union, "by_queue": by_queue}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/current-operators")
def get_current_operators(
    partner_uuid: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Текущий статус операторов проекта — берёт логины прямо из Naumen."""
    # Local name overrides (our DB has richer names)
    employees = db.query(Employee).filter(
        Employee.project_uuid == partner_uuid,
        Employee.naumen_login.isnot(None),
    ).all()
    local_names = {e.naumen_login: e.full_name for e in employees}
    try:
        overrides = _build_overrides(db)
        work_statuses, offline_statuses = _status_sets(db, partner_uuid)
        rows = naumen.get_current_operators_for_project(
            partner_uuid, overrides, work_statuses, offline_statuses,
        )
        # Очереди оператора (по обработанным звонкам за неделю) — для показа
        # рядом с ФИО в Мониторинге и фильтрации по очередям. Кэшируется на 5 мин.
        try:
            queues_map = _operator_queues_cached(partner_uuid, overrides)
        except Exception:
            queues_map = {}
        for r in rows:
            local = local_names.get(r.get("login"))
            if local:
                r["employee_name"] = local
            r["queues"] = queues_map.get(r.get("login"), [])
        return {"data": rows, "total_logins": len(rows)}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/current-operators-outbound")
def get_current_operators_outbound(
    partner_uuid: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Текущий статус операторов ИСХОДЯЩИХ линий проекта (для Мониторинга, линия «Исход»)."""
    employees = db.query(Employee).filter(
        Employee.project_uuid == partner_uuid,
        Employee.naumen_login.isnot(None),
    ).all()
    local_names = {e.naumen_login: e.full_name for e in employees}
    try:
        overrides = _build_overrides(db)
        work_statuses, offline_statuses = _status_sets(db, partner_uuid)
        rows = naumen.get_current_operators_outbound(
            partner_uuid, overrides, work_statuses, offline_statuses,
        )
        try:
            projects_map = _operator_outbound_projects_cached(partner_uuid, overrides)
        except Exception:
            projects_map = {}
        for r in rows:
            local = local_names.get(r.get("login"))
            if local:
                r["employee_name"] = local
            r["queues"] = projects_map.get(r.get("login"), [])
        return {"data": rows, "total_logins": len(rows)}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/recent-stats-outbound")
def get_recent_stats_outbound_ep(
    partner_uuid: str,
    window_min: int = Query(1440, ge=1, le=1440),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Live-статистика исходящего обзвона за последние window_min минут: по
    операторам и по результату попыток (Мониторинг → Статистика, линия «Исход»)."""
    check_project_access(partner_uuid, current_user, db)
    try:
        return naumen.get_recent_stats_outbound(partner_uuid, int(window_min), _build_overrides(db))
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/outbound-projects")
def outbound_projects_ep(
    partner_uuid: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Список исходящих подпроектов («очередей» обзвона) партнёра + флаг hidden
    (скрытые в настройках — по queue_settings, ключ = uuid подпроекта)."""
    check_project_access(partner_uuid, current_user, db)
    try:
        projects = _outbound_projects_list_cached(partner_uuid, _build_overrides(db))
        # Настройки подпроектов храним в queue_settings с ключом "out:<uuid>"
        # (направление Вход/Исход + видимость), чтобы не путать их со входящими.
        settings = {
            r.queue_name: r
            for r in db.query(QueueSetting).filter(QueueSetting.partner_uuid == partner_uuid).all()
        }
        for p in projects:
            ov = settings.get(f"out:{p.get('project_uuid')}")
            # По умолчанию исходящий подпроект = «Исход», видим.
            p["show_in"] = bool(ov.show_in) if ov and ov.show_in is not None else False
            p["show_out"] = bool(ov.show_out) if ov and ov.show_out is not None else True
            p["hidden"] = bool(ov.hidden) if ov and ov.hidden is not None else False
        return {"data": projects}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/outbound-summary")
def outbound_summary_ep(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    project_ids: Optional[list[str]] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Сводка исходящего обзвона за период (раздел «Аналитика (Исход)»):
    попытки, кейсы, контакты, contact rate, ср. разговор, попыток на кейс,
    распределение по результату, динамика по дням и разрез по подпроектам.
    project_ids — фильтр по конкретным исходящим подпроектам."""
    check_project_access(partner_uuid, current_user, db)
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        return naumen.get_outbound_summary(partner_uuid, str(begin), _exclusive_end(end), project_ids or None, _build_overrides(db))
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/outbound-operators")
def outbound_operators_ep(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    project_ids: Optional[list[str]] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Нагрузка операторов на обзвоне за период (с фильтром по подпроектам)."""
    check_project_access(partner_uuid, current_user, db)
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        return {"data": naumen.get_outbound_operator_load(partner_uuid, str(begin), _exclusive_end(end), project_ids or None, _build_overrides(db))}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/outbound-load")
def outbound_load_ep(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    project_ids: Optional[list[str]] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Нагрузка обзвона по часам суток за период (с фильтром по подпроектам)."""
    check_project_access(partner_uuid, current_user, db)
    if end < begin:
        raise HTTPException(400, detail="end не может быть раньше begin")
    try:
        return {"data": naumen.get_outbound_load_by_hour(partner_uuid, str(begin), _exclusive_end(end), project_ids or None, _build_overrides(db))}
    except Exception as e:
        raise HTTPException(503, detail=str(e))


@router.get("/naumen-employees")
def get_naumen_employees(
    partner_uuid: str,
    begin: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    try:
        data = naumen.get_naumen_employees(partner_uuid, str(begin), _exclusive_end(end), _build_overrides(db))
        return {"data": data}
    except Exception as e:
        raise HTTPException(503, detail=str(e))
