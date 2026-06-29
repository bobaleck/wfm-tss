from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from typing import List, Optional
from datetime import date

from app.core.database import get_db
from app.models.schedule import Schedule, Shift, Absence
from app.models.employee import Employee
from datetime import datetime
from app.schemas.schedule import (
    ScheduleCreate, ScheduleUpdate, ScheduleOut,
    ShiftCreate, ShiftUpdate, ShiftOut, ShiftConfirm,
    AbsenceCreate, AbsenceUpdate, AbsenceOut,
)
from app.api.deps import (
    get_current_user, require_manager, check_project_access,
    resolve_project_scope, accessible_project_uuids,
)

router = APIRouter()


# ─── Шаблоны расписаний ───────────────────────────────────────────────────────

@router.get("", response_model=List[ScheduleOut])
def list_schedules(project_uuid: Optional[str] = None, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    scope = resolve_project_scope(project_uuid, current_user, db)
    q = db.query(Schedule)
    if scope is not None:
        # Графики строго проектные — индивидуальны для каждого проекта.
        q = q.filter(Schedule.project_uuid.in_(scope))
    return q.order_by(Schedule.name).all()


@router.post("", response_model=ScheduleOut, status_code=201)
def create_schedule(body: ScheduleCreate, db: Session = Depends(get_db), current_user=Depends(require_manager)):
    if getattr(body, "project_uuid", None):
        check_project_access(body.project_uuid, current_user, db)
    s = Schedule(**body.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


@router.put("/{schedule_id}", response_model=ScheduleOut)
def update_schedule(schedule_id: int, body: ScheduleUpdate,
                    db: Session = Depends(get_db), current_user=Depends(require_manager)):
    s = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not s:
        raise HTTPException(404, "Не найдено")
    if s.project_uuid:
        check_project_access(s.project_uuid, current_user, db)
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return s


@router.delete("/{schedule_id}", status_code=204)
def delete_schedule(schedule_id: int, db: Session = Depends(get_db), current_user=Depends(require_manager)):
    s = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not s:
        raise HTTPException(404, "Не найдено")
    if s.project_uuid:
        check_project_access(s.project_uuid, current_user, db)
    db.delete(s)
    db.commit()


# ─── Смены сотрудников ────────────────────────────────────────────────────────

@router.get("/shifts", response_model=List[ShiftOut])
def list_shifts(
    employee_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    project_uuid: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    scope = resolve_project_scope(project_uuid, current_user, db)
    q = db.query(Shift).join(Employee)
    if employee_id:
        q = q.filter(Shift.employee_id == employee_id)
    if date_from:
        q = q.filter(Shift.shift_date >= date_from)
    if date_to:
        q = q.filter(Shift.shift_date <= date_to)
    if scope is not None:
        q = q.filter(Employee.project_uuid.in_(scope))
    shifts = q.order_by(Shift.shift_date, Shift.start_time).all()
    result = []
    for sh in shifts:
        out = ShiftOut.model_validate(sh).model_dump()
        out["employee_name"] = sh.employee.full_name if sh.employee else None
        out["schedule_name"] = sh.schedule.name if sh.schedule else None
        out["team_id"] = sh.employee.team_id if sh.employee else None
        out["team_name"] = sh.employee.team.name if sh.employee and sh.employee.team else None
        result.append(out)
    return result


@router.post("/shifts", response_model=ShiftOut, status_code=201)
def create_shift(body: ShiftCreate, db: Session = Depends(get_db), current_user=Depends(require_manager)):
    emp = db.query(Employee).filter(Employee.id == body.employee_id).first()
    if emp and emp.project_uuid:
        check_project_access(emp.project_uuid, current_user, db)
    shift = Shift(**body.model_dump())
    db.add(shift)
    db.commit()
    db.refresh(shift)
    out = ShiftOut.model_validate(shift).model_dump()
    out["employee_name"] = shift.employee.full_name if shift.employee else None
    out["schedule_name"] = shift.schedule.name if shift.schedule else None
    return out


@router.put("/shifts/{shift_id}", response_model=ShiftOut)
def update_shift(shift_id: int, body: ShiftUpdate,
                 db: Session = Depends(get_db), current_user=Depends(require_manager)):
    shift = db.query(Shift).filter(Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(404, "Не найдена")
    if shift.employee and shift.employee.project_uuid:
        check_project_access(shift.employee.project_uuid, current_user, db)
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(shift, k, v)
    db.commit()
    db.refresh(shift)
    out = ShiftOut.model_validate(shift).model_dump()
    out["employee_name"] = shift.employee.full_name if shift.employee else None
    out["schedule_name"] = shift.schedule.name if shift.schedule else None
    return out


@router.delete("/shifts/{shift_id}", status_code=204)
def delete_shift(shift_id: int, db: Session = Depends(get_db), current_user=Depends(require_manager)):
    shift = db.query(Shift).filter(Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(404, "Не найдена")
    if shift.employee and shift.employee.project_uuid:
        check_project_access(shift.employee.project_uuid, current_user, db)
    db.delete(shift)
    db.commit()


@router.post("/shifts/{shift_id}/confirm")
def confirm_shift(
    shift_id: int,
    body: ShiftConfirm,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager),
):
    """Подтвердить смену и/или указать фактически отработанные часы."""
    shift = db.query(Shift).filter(Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(404, "Смена не найдена")
    if shift.employee and shift.employee.project_uuid:
        check_project_access(shift.employee.project_uuid, current_user, db)
    shift.actual_start_time = body.actual_start_time or shift.start_time
    shift.actual_end_time = body.actual_end_time or shift.end_time
    shift.actual_hours_worked = body.actual_hours_worked
    shift.needs_review = False
    shift.status = "completed"
    shift.reconciled_at = datetime.utcnow()
    db.commit()
    db.refresh(shift)
    out = ShiftOut.model_validate(shift).model_dump()
    out["employee_name"] = shift.employee.full_name if shift.employee else None
    out["schedule_name"] = shift.schedule.name if shift.schedule else None
    return out


@router.post("/shifts/reconcile")
def reconcile_shifts(
    reconcile_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager),
):
    """
    Сверка плановых смен с данными Naumen за указанную дату (по умолчанию — вчера).
    Вычисляет фактически отработанные часы из status_changes и помечает расхождения.
    """
    from datetime import date as date_type, timedelta
    from app.models.audit import IntegrationSettings
    import app.services.naumen_db as naumen

    target_date = reconcile_date or (date_type.today() - timedelta(days=1))

    s = db.query(IntegrationSettings).first()
    overrides = {
        "host": s.db_host, "database": s.db_name,
        "user": s.db_user, "password": s.db_password, "port": s.db_port,
    } if s and s.db_host else None

    # Проектная изоляция: сверяем только смены доступных пользователю проектов
    # (Админ/Аналитик — все). Менеджер не может пересчитать чужие проекты.
    scope = accessible_project_uuids(current_user, db)
    sq = db.query(Shift).join(Employee).filter(
        Shift.shift_date == target_date,
        Shift.status.in_(["planned", "confirmed"]),
    )
    if scope is not None:
        sq = sq.filter(Employee.project_uuid.in_(scope))
    shifts = sq.all()

    updated, flagged, skipped = 0, 0, 0
    for shift in shifts:
        emp = shift.employee
        if not emp or not emp.naumen_login:
            skipped += 1
            continue
        try:
            work_sec = naumen.get_operator_day_seconds(emp.naumen_login, str(target_date), overrides)
        except Exception:
            skipped += 1
            continue

        actual_hours = round(work_sec / 3600, 2)
        shift.actual_hours_worked = str(actual_hours)
        shift.reconciled_at = datetime.utcnow()

        planned_hours = 0.0
        if shift.start_time and shift.end_time:
            try:
                from datetime import datetime as dt
                fmt = "%Y-%m-%dT%H:%M"
                s_dt = dt.fromisoformat(shift.start_time[:16])
                e_dt = dt.fromisoformat(shift.end_time[:16])
                planned_hours = (e_dt - s_dt).total_seconds() / 3600
            except Exception:
                pass

        if actual_hours > 0:
            shift.status = "completed"
            if planned_hours > 0 and abs(actual_hours - planned_hours) > 1.0:
                shift.needs_review = True
                flagged += 1
            else:
                shift.needs_review = False
                updated += 1
        else:
            skipped += 1

    db.commit()
    return {"ok": True, "date": str(target_date), "updated": updated, "flagged": flagged, "skipped": skipped}


@router.get("/shifts/export.xlsx")
def export_shifts_xlsx(
    project_uuid: Optional[str] = None,
    date_from: date = Query(...),
    date_to: date = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Выгрузка графика смен в Excel по шаблону «График работы.xlsx»:
    строки — сотрудники (сгруппированы по команде), столбцы — по 2 на каждый день
    (дата+день недели в шапке, в ячейке — время смены и плановые часы)."""
    import io
    from datetime import timedelta, datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    if date_to < date_from:
        raise HTTPException(400, detail="date_to не может быть раньше date_from")

    days = []
    d = date_from
    while d <= date_to:
        days.append(d)
        d += timedelta(days=1)

    scope = resolve_project_scope(project_uuid, current_user, db)
    sq = db.query(Shift).join(Employee).filter(
        Shift.shift_date >= date_from, Shift.shift_date <= date_to,
    )
    if scope is not None:
        sq = sq.filter(Employee.project_uuid.in_(scope))
    by_emp_date = {(sh.employee_id, sh.shift_date): sh for sh in sq.all()}

    eq = db.query(Employee)
    if scope is not None:
        eq = eq.filter(Employee.project_uuid.in_(scope))
    employees = [e for e in eq.all() if e.employment_status != "fired"]

    def team_name(e):
        return e.team.name if e.team else "Без команды"
    employees.sort(key=lambda e: (team_name(e), e.full_name or ""))

    WEEKDAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="434343")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    name_fill = PatternFill("solid", fgColor="B6D7A8")
    day_fill = PatternFill("solid", fgColor="D9EAD3")
    night_fill = PatternFill("solid", fgColor="D9D2E9")
    we_fill = PatternFill("solid", fgColor="EFEFEF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    fixed = ["Команда", "Логин", "ФИО", "Telegram", "Должность", "График", "Тип"]
    NF = len(fixed)
    for i, h in enumerate(fixed, start=1):
        c = ws.cell(row=1, column=i, value=h)
        ws.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

    for di, day in enumerate(days):
        col = NF + 1 + di * 2
        c1 = ws.cell(row=1, column=col, value=day.strftime("%d.%m"))
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        c2 = ws.cell(row=2, column=col, value=WEEKDAYS[day.weekday()])
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        s1 = ws.cell(row=3, column=col, value="Смена")
        s2 = ws.cell(row=3, column=col + 1, value="Ч")
        for cc in (c1, c2, s1, s2):
            cc.fill = hdr_fill; cc.font = hdr_font; cc.alignment = center; cc.border = border

    widths = [26, 16, 30, 18, 16, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for di in range(len(days)):
        col = NF + 1 + di * 2
        ws.column_dimensions[get_column_letter(col)].width = 11
        ws.column_dimensions[get_column_letter(col + 1)].width = 6

    r = 4
    for e in employees:
        ws.cell(row=r, column=1, value=team_name(e)).alignment = left
        ws.cell(row=r, column=2, value=e.naumen_login or "")
        nc = ws.cell(row=r, column=3, value=e.full_name or ""); nc.fill = name_fill
        ws.cell(row=r, column=4, value=e.email or "")
        ws.cell(row=r, column=5, value=e.position or "")
        ws.cell(row=r, column=6, value=e.preferred_schedule or "")
        ws.cell(row=r, column=7, value="")
        for i in range(1, NF + 1):
            ws.cell(row=r, column=i).border = border
        for di, day in enumerate(days):
            col = NF + 1 + di * 2
            sh = by_emp_date.get((e.id, day))
            cell = ws.cell(row=r, column=col)
            hcell = ws.cell(row=r, column=col + 1)
            is_we = day.weekday() >= 5
            if sh and sh.start_time and sh.end_time:
                st = sh.start_time[11:16]; en = sh.end_time[11:16]
                cell.value = f"{st}-{en}"
                try:
                    a = _dt.fromisoformat(sh.start_time[:16]); b = _dt.fromisoformat(sh.end_time[:16])
                    hrs = (b - a).total_seconds() / 3600 - (sh.lunch_minutes or 0) / 60
                    hcell.value = round(hrs, 1)
                except Exception:
                    pass
                cell.fill = night_fill if en <= st else day_fill
            elif is_we:
                cell.fill = we_fill; hcell.fill = we_fill
            cell.alignment = center; cell.border = border
            hcell.alignment = center; hcell.border = border
        r += 1

    ws.freeze_panes = "H4"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"shifts_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── Отсутствия ───────────────────────────────────────────────────────────────

@router.get("/absences", response_model=List[AbsenceOut])
def list_absences(
    employee_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    absence_type: Optional[str] = None,
    project_uuid: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    scope = resolve_project_scope(project_uuid, current_user, db)
    q = db.query(Absence)
    if scope is not None:
        q = q.join(Employee).filter(Employee.project_uuid.in_(scope))
    if employee_id:
        q = q.filter(Absence.employee_id == employee_id)
    if date_from:
        q = q.filter(Absence.end_date >= date_from)
    if date_to:
        q = q.filter(Absence.start_date <= date_to)
    if absence_type:
        q = q.filter(Absence.absence_type == absence_type)
    absences = q.order_by(Absence.start_date).all()
    result = []
    for ab in absences:
        out = AbsenceOut.model_validate(ab).model_dump()
        out["employee_name"] = ab.employee.full_name if ab.employee else None
        result.append(out)
    return result


@router.post("/absences", response_model=AbsenceOut, status_code=201)
def create_absence(body: AbsenceCreate, db: Session = Depends(get_db), current_user=Depends(require_manager)):
    emp = db.query(Employee).filter(Employee.id == body.employee_id).first()
    if emp and emp.project_uuid:
        check_project_access(emp.project_uuid, current_user, db)
    ab = Absence(**body.model_dump())
    db.add(ab)
    db.commit()
    db.refresh(ab)
    out = AbsenceOut.model_validate(ab).model_dump()
    out["employee_name"] = ab.employee.full_name if ab.employee else None
    return out


@router.put("/absences/{absence_id}", response_model=AbsenceOut)
def update_absence(absence_id: int, body: AbsenceUpdate,
                   db: Session = Depends(get_db), current_user=Depends(require_manager)):
    ab = db.query(Absence).filter(Absence.id == absence_id).first()
    if not ab:
        raise HTTPException(404, "Не найдено")
    if ab.employee and ab.employee.project_uuid:
        check_project_access(ab.employee.project_uuid, current_user, db)
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(ab, k, v)
    db.commit()
    db.refresh(ab)
    out = AbsenceOut.model_validate(ab).model_dump()
    out["employee_name"] = ab.employee.full_name if ab.employee else None
    return out


@router.delete("/absences/{absence_id}", status_code=204)
def delete_absence(absence_id: int, db: Session = Depends(get_db), current_user=Depends(require_manager)):
    ab = db.query(Absence).filter(Absence.id == absence_id).first()
    if not ab:
        raise HTTPException(404, "Не найдено")
    if ab.employee and ab.employee.project_uuid:
        check_project_access(ab.employee.project_uuid, current_user, db)
    db.delete(ab)
    db.commit()
